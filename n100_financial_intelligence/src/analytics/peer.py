import os
import sqlite3
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def get_peer_percentiles(db_path):
    conn = sqlite3.connect(db_path)
    
    # Load peer groups
    pg_df = pd.read_sql_query("SELECT * FROM peer_groups", conn)
    
    # Load ratios & companies
    ratios_df = pd.read_sql_query("""
        SELECT fr.*, c.roce_percentage
        FROM financial_ratios fr
        JOIN companies c ON fr.company_id = c.id
    """, conn)
    
    conn.close()
    
    # Unique groups
    groups = pg_df['peer_group_name'].unique()
    
    # Metrics to rank (10 metrics)
    metrics_map = {
        "ROE": "return_on_equity_pct",
        "ROCE": "roce_percentage",
        "Net Profit Margin": "net_profit_margin_pct",
        "D/E": "debt_to_equity",
        "FCF": "free_cash_flow_cr",
        "PAT CAGR 5yr": "pat_cagr_5yr",
        "Revenue CAGR 5yr": "revenue_cagr_5yr",
        "EPS CAGR 5yr": "eps_cagr_5yr",
        "Interest Coverage": "interest_coverage",
        "Asset Turnover": "asset_turnover"
    }
    
    records = []
    
    # Calculate for each year
    all_years = ratios_df['year'].unique()
    
    for yr in all_years:
        yr_ratios = ratios_df[ratios_df['year'] == yr]
        
        for g_name in groups:
            # Get companies in group
            g_comps = pg_df[pg_df['peer_group_name'] == g_name]['company_id'].unique()
            g_ratios = yr_ratios[yr_ratios['company_id'].isin(g_comps)]
            
            if len(g_ratios) == 0:
                continue
                
            for metric_label, col in metrics_map.items():
                # Extract values
                series = g_ratios.set_index('company_id')[col]
                # Drop NaN for ranking
                valid_series = series.dropna()
                
                if len(valid_series) == 0:
                    continue
                    
                # Rank (pct=True returns 0 to 1)
                ranks = valid_series.rank(pct=True, method='min')
                
                # Invert for D/E
                if metric_label == "D/E":
                    ranks = 1.0 - ranks
                    
                for comp_id, val in valid_series.items():
                    pct_rank = ranks.loc[comp_id]
                    records.append({
                        "company_id": comp_id,
                        "peer_group_name": g_name,
                        "metric": metric_label,
                        "value": float(val),
                        "percentile_rank": float(pct_rank),
                        "year": int(yr)
                    })
                    
    # Save to SQLite table peer_percentiles
    conn = sqlite3.connect(db_path)
    conn.execute("DELETE FROM peer_percentiles")
    conn.commit()
    
    if records:
        out_df = pd.DataFrame(records)
        out_df.to_sql("peer_percentiles", conn, if_exists="append", index=False)
        conn.commit()
        print(f"Loaded {len(out_df)} rows into peer_percentiles table.")
    conn.close()
    
    return records

def check_company_peer_group(db_path, company_id):
    """If company not in any peer group, return 'No peer group assigned' message."""
    conn = sqlite3.connect(db_path)
    res = conn.execute("SELECT COUNT(*) FROM peer_groups WHERE company_id = ?", (company_id,)).fetchone()[0]
    conn.close()
    if res == 0:
        return "No peer group assigned"
    return "Assigned"

def generate_radar_charts(db_path):
    """
    Generates radar/polar chart for each company in a peer group.
    8 axes: ROE, ROCE, NPM, D/E, FCF score, PAT CAGR 5yr, Revenue CAGR 5yr, Composite Score.
    """
    conn = sqlite3.connect(db_path)
    pg_df = pd.read_sql_query("SELECT * FROM peer_groups", conn)
    
    # We want latest metrics for radar charts (usually 2024)
    # Join with latest ratios
    df_latest = pd.read_sql_query("""
        WITH LatestRatio AS (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY company_id ORDER BY year DESC) as rn
            FROM financial_ratios
        )
        SELECT lr.*, c.roce_percentage, s.broad_sector
        FROM LatestRatio lr
        JOIN companies c ON lr.company_id = c.id
        LEFT JOIN sectors s ON lr.company_id = s.company_id
        WHERE lr.rn = 1
    """, conn)
    
    conn.close()
    
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "reports", "radar_charts")
    os.makedirs(output_dir, exist_ok=True)
    
    # 8 axes mapping
    axes = ["ROE", "ROCE", "NPM", "D/E", "FCF", "PAT CAGR", "Rev CAGR", "Composite"]
    col_mapping = {
        "ROE": "return_on_equity_pct",
        "ROCE": "roce_percentage",
        "NPM": "net_profit_margin_pct",
        "D/E": "debt_to_equity",
        "FCF": "free_cash_flow_cr",
        "PAT CAGR": "pat_cagr_5yr",
        "Rev CAGR": "revenue_cagr_5yr",
        "Composite": "composite_quality_score"
    }
    
    # Normalize latest ratios from 0-100 to display on radar chart
    norm_df = df_latest.copy()
    for col in col_mapping.values():
        p10 = norm_df[col].quantile(0.1)
        p90 = norm_df[col].quantile(0.9)
        diff = p90 - p10 if p90 > p10 else 1.0
        norm_df[col + "_norm"] = (norm_df[col].clip(lower=p10, upper=p90) - p10) / diff * 100.0
        
    for idx, row in norm_df.iterrows():
        comp_id = row['company_id']
        
        # Check if company has peer group
        in_group = pg_df[pg_df['company_id'] == comp_id]
        if len(in_group) == 0:
            # Standalone chart
            # Just plot company value as bar or single metric
            continue
            
        g_name = in_group.iloc[0]['peer_group_name']
        g_comps = pg_df[pg_df['peer_group_name'] == g_name]['company_id'].unique()
        group_df = norm_df[norm_df['company_id'].isin(g_comps)]
        
        # Company values
        comp_vals = []
        group_avg_vals = []
        
        for axis in axes:
            col_norm = col_mapping[axis] + "_norm"
            comp_vals.append(float(row[col_norm]) if pd.notna(row[col_norm]) else 50.0)
            group_avg_vals.append(float(group_df[col_norm].mean()) if pd.notna(group_df[col_norm].mean()) else 50.0)
            
        # Radar plot styling
        num_vars = len(axes)
        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
        
        # Close the loop
        comp_vals += comp_vals[:1]
        group_avg_vals += group_avg_vals[:1]
        angles += angles[:1]
        
        fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
        
        # Draw axes
        plt.xticks(angles[:-1], axes, color='grey', size=10)
        
        # Draw ylabels
        ax.set_rlabel_position(0)
        plt.yticks([25, 50, 75, 100], ["25", "50", "75", "100"], color="grey", size=8)
        plt.ylim(0, 100)
        
        # Plot company
        ax.plot(angles, comp_vals, linewidth=2, linestyle='solid', label=comp_id, color="#1F4E79")
        ax.fill(angles, comp_vals, color="#1F4E79", alpha=0.25)
        
        # Plot group average
        ax.plot(angles, group_avg_vals, linewidth=1.5, linestyle='dashed', label=f"{g_name} Avg", color="#D95F02")
        
        plt.title(f"{comp_id} Performance Radar vs {g_name}", size=12, color="#1F4E79", y=1.1)
        plt.legend(loc='upper right', bbox_to_anchor=(0.1, 0.1))
        
        # Save PNG
        chart_path = os.path.join(output_dir, f"{comp_id}_radar.png")
        plt.savefig(chart_path, dpi=100, bbox_inches='tight')
        plt.close()
        
    print(f"Generated radar charts in {output_dir}")

def generate_peer_comparison_report(db_path):
    conn = sqlite3.connect(db_path)
    pg_df = pd.read_sql_query("SELECT * FROM peer_groups", conn)
    
    # Ratios
    ratios_df = pd.read_sql_query("""
        WITH LatestRatio AS (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY company_id ORDER BY year DESC) as rn
            FROM financial_ratios
        )
        SELECT lr.*, c.company_name, c.roce_percentage
        FROM LatestRatio lr
        JOIN companies c ON lr.company_id = c.id
        WHERE lr.rn = 1
    """, conn)
    
    # Percentiles
    pct_df = pd.read_sql_query("""
        WITH LatestPct AS (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY company_id, metric ORDER BY year DESC) as rn
            FROM peer_percentiles
        )
        SELECT * FROM LatestPct WHERE rn = 1
    """, conn)
    
    conn.close()
    
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    output_path = os.path.join(base_dir, "output", "peer_comparison.xlsx")
    
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    font_body = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    fill_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Soft green for >= 75th pct
    fill_mid = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft yellow for 25th-75th
    fill_low = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Soft red for <= 25th
    fill_bench = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Gold/amber for benchmark
    
    # 10 metrics we computed percentiles for
    metrics = ["ROE", "ROCE", "Net Profit Margin", "D/E", "FCF", "PAT CAGR 5yr", "Revenue CAGR 5yr", "EPS CAGR 5yr", "Interest Coverage", "Asset Turnover"]
    
    for g_name in pg_df['peer_group_name'].unique():
        ws = wb.create_sheet(title=g_name)
        
        # Build headers
        headers = ["company_id", "company_name"]
        for m in metrics:
            headers.append(f"{m} Value")
            headers.append(f"{m} Percentile")
            
        ws.append(headers)
        
        # Write rows
        g_comps = pg_df[pg_df['peer_group_name'] == g_name]
        
        row_idx = 2
        for _, gc in g_comps.iterrows():
            comp_id = gc['company_id']
            is_bench = gc['is_benchmark']
            
            comp_r = ratios_df[ratios_df['company_id'] == comp_id]
            if len(comp_r) == 0:
                continue
                
            r_row = comp_r.iloc[0]
            comp_name = r_row['company_name']
            
            row_data = [comp_id, comp_name]
            
            for m in metrics:
                # Value
                val = None
                if m == "ROE": val = r_row['return_on_equity_pct']
                elif m == "ROCE": val = r_row['roce_percentage']
                elif m == "Net Profit Margin": val = r_row['net_profit_margin_pct']
                elif m == "D/E": val = r_row['debt_to_equity']
                elif m == "FCF": val = r_row['free_cash_flow_cr']
                elif m == "PAT CAGR 5yr": val = r_row['pat_cagr_5yr']
                elif m == "Revenue CAGR 5yr": val = r_row['revenue_cagr_5yr']
                elif m == "EPS CAGR 5yr": val = r_row['eps_cagr_5yr']
                elif m == "Interest Coverage": val = r_row['interest_coverage']
                elif m == "Asset Turnover": val = r_row['asset_turnover']
                
                # Percentile
                pct_val = pct_df[(pct_df['company_id'] == comp_id) & (pct_df['metric'] == m)]
                pct_rank = pct_val.iloc[0]['percentile_rank'] if len(pct_val) > 0 else None
                
                row_data.append(val)
                row_data.append(pct_rank)
                
            ws.append(row_data)
            
            # Format row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.border = thin_border
                
                if col_idx in [1, 2]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
                    
                # Format Percentile rank cells & color
                if col_idx > 2 and col_idx % 2 == 0: # Even column indexes are Percentile ranks
                    pct_val = cell.value
                    if isinstance(pct_val, (int, float)):
                        cell.number_format = '0.0%'
                        if pct_val >= 0.75:
                            cell.fill = fill_high
                        elif pct_val <= 0.25:
                            cell.fill = fill_low
                        else:
                            cell.fill = fill_mid
                            
                # Apply benchmark coloring
                if is_bench:
                    if col_idx <= 2: # Keep ticker/name gold
                        cell.fill = fill_bench
                        cell.font = font_bold
            row_idx += 1
            
        # Add median row at the bottom
        median_row = ["Median", "Peer Group Median"]
        for col_idx in range(3, len(headers) + 1):
            vals = []
            for r in range(2, row_idx):
                val = ws.cell(row=r, column=col_idx).value
                if isinstance(val, (int, float)):
                    vals.append(val)
            median_row.append(np.median(vals) if vals else None)
            
        ws.append(median_row)
        
        # Format median row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_bold
            cell.border = thin_border
            if col_idx in [1, 2]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if col_idx % 2 == 0:
                    cell.number_format = '0.0%'
                    
        # Apply header styling
        for col_idx in range(1, len(headers) + 1):
            h_cell = ws.cell(row=1, column=col_idx)
            h_cell.font = font_header
            h_cell.fill = fill_header
            h_cell.alignment = align_center
            h_cell.border = thin_border
            
        # Column widths auto fit
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = ws.cell(row=1, column=col[0].column).column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Peer comparison report generated at {output_path}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    db_path = os.path.join(base_dir, "data", "nifty100.db")
    get_peer_percentiles(db_path)
    generate_radar_charts(db_path)
    generate_peer_comparison_report(db_path)

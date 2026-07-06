import os
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from engine import ScreenerEngine

def get_3yr_rev_cagr_and_de_trend(db_path):
    """
    Computes 3-year Revenue CAGR and D/E trend (is latest D/E < previous year D/E)
    for all valid companies.
    """
    conn = sqlite3.connect(db_path)
    
    # Sales data
    pl_df = pd.read_sql_query("SELECT company_id, year, sales FROM profitandloss", conn)
    # D/E data
    fr_df = pd.read_sql_query("SELECT company_id, year, debt_to_equity FROM financial_ratios", conn)
    
    conn.close()
    
    results = {}
    
    # 3-year Revenue CAGR
    for comp in pl_df['company_id'].unique():
        comp_pl = pl_df[pl_df['company_id'] == comp].sort_values(by='year')
        if len(comp_pl) >= 4:
            latest_sales = comp_pl.iloc[-1]['sales']
            start_sales = comp_pl.iloc[-4]['sales']
            if start_sales > 0 and latest_sales > 0:
                cagr_3yr = ((latest_sales / start_sales) ** (1.0 / 3) - 1.0) * 100.0
            else:
                cagr_3yr = None
        else:
            cagr_3yr = None
            
        # D/E trend
        comp_fr = fr_df[fr_df['company_id'] == comp].sort_values(by='year')
        if len(comp_fr) >= 2:
            latest_de = comp_fr.iloc[-1]['debt_to_equity']
            prev_de = comp_fr.iloc[-2]['debt_to_equity']
            de_declining = (latest_de < prev_de) if (latest_de is not None and prev_de is not None) else False
        else:
            de_declining = False
            
        results[comp] = {
            "rev_cagr_3yr": cagr_3yr,
            "de_declining": de_declining
        }
    return results

def compute_sector_relative_scores(df):
    """
    Computes winsorized sector-relative composite scores on 0-100 scale:
    - 35% Profitability (ROE 15% + ROCE 10% + NPM 10%)
    - 30% Cash Quality (FCF CAGR 15% + CFO/PAT ratio 10% + FCF positive flag 5%)
    - 20% Growth (Revenue CAGR 10% + PAT CAGR 10%)
    - 15% Leverage (D/E score 10% + ICR score 5%)
    """
    scored_df = df.copy()
    
    # Temporary copy of metrics needed
    metrics = {
        "roe": ("return_on_equity_pct", "max", 0.15),
        "roce": ("roce_percentage", "max", 0.10), # ROCE fallback
        "npm": ("net_profit_margin_pct", "max", 0.10),
        "fcf_cagr": ("pat_cagr_5yr", "max", 0.15), # FCF CAGR fallback to PAT CAGR
        "cfo_pat": ("cash_from_operations_cr", "max", 0.10), # raw CFO metric fallback
        "rev_cagr": ("revenue_cagr_5yr", "max", 0.10),
        "pat_cagr": ("pat_cagr_5yr", "max", 0.10),
        "de": ("debt_to_equity", "min", 0.10),
        "icr": ("interest_coverage", "max", 0.05)
    }
    
    for key, (col, direction, weight) in metrics.items():
        score_col = key + "_score"
        scored_df[score_col] = 50.0 # fallback
        
        # Calculate within each broad sector
        for sector in scored_df['broad_sector'].unique():
            sector_mask = scored_df['broad_sector'] == sector
            sub_df = scored_df[sector_mask]
            
            vals = pd.to_numeric(sub_df[col], errors='coerce').dropna()
            if len(vals) < 3:
                continue
                
            p10 = vals.quantile(0.10)
            p90 = vals.quantile(0.90)
            diff = p90 - p10 if p90 > p10 else 1.0
            
            clipped = sub_df[col].clip(lower=p10, upper=p90)
            
            if direction == "max":
                scores = (clipped - p10) / diff * 100.0
            else: # min is better (like D/E)
                scores = (p90 - clipped) / diff * 100.0
                
            scored_df.loc[sector_mask, score_col] = scores.fillna(50.0)
            
    # CFO/PAT ratio and FCF positive flag scores
    # FCF flag: 100 if FCF > 0 else 0
    fcf_flag_score = scored_df['free_cash_flow_cr'].apply(lambda x: 100.0 if x is not None and x > 0 else 0.0)
    
    # Calculate composite score
    scored_df['sector_composite_quality_score'] = (
        0.15 * scored_df['roe_score'] +
        0.10 * scored_df['roce_score'] +
        0.10 * scored_df['npm_score'] +
        0.15 * scored_df['fcf_cagr_score'] +
        0.10 * scored_df['cfo_pat_score'] +
        0.05 * fcf_flag_score +
        0.10 * scored_df['rev_cagr_score'] +
        0.10 * scored_df['pat_cagr_score'] +
        0.10 * scored_df['de_score'] +
        0.05 * scored_df['icr_score']
    )
    
    # Drop temp score columns
    temp_cols = [k + "_score" for k in metrics.keys()]
    scored_df = scored_df.drop(columns=temp_cols)
    return scored_df

def apply_excel_styling(ws, preset_name):
    """Applies premium styling and threshold-based color coding to cells."""
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Classic corporate dark blue
    
    font_body = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Format headers
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Styles for threshold coloring
    fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
    fill_fail = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
    
    # Format cells
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = thin_border
            
            # Alignments based on type
            if col_idx in [1, 2]: # Ticker, name
                cell.alignment = align_left
            elif col_idx in [3, 4]: # broad_sector, sub_sector
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                
            # Number formatting
            val = cell.value
            header = ws.cell(row=1, column=col_idx).value
            
            # Number formats
            if isinstance(val, (int, float)):
                if "ratio" in header.lower() or "multiple" in header.lower() or "debt_to_equity" in header:
                    cell.number_format = '0.00'
                elif "percentage" in header.lower() or "pct" in header.lower() or "cagr" in header.lower() or "roe" in header.lower() or "roce" in header.lower() or "npm" in header.lower() or "opm" in header.lower():
                    cell.number_format = '0.0%' if val <= 1.0 else '0.0'
                elif "score" in header.lower():
                    cell.number_format = '0.0'
                else: # Cr / Volume
                    cell.number_format = '#,##0.0'
                    
            # Color-code threshold criteria based on preset rules
            if preset_name == "Quality Compounder":
                if header == "return_on_equity_pct" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 15.0 else fill_fail
                elif header == "debt_to_equity" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val <= 1.0 else fill_fail
                elif header == "free_cash_flow_cr" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val > 0 else fill_fail
                elif header == "revenue_cagr_5yr" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 10.0 else fill_fail
                    
            elif preset_name == "Value Pick":
                if header == "pe_ratio" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val <= 30.0 else fill_fail
                elif header == "pb_ratio" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val <= 5.0 else fill_fail
                elif header == "debt_to_equity" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val <= 2.0 else fill_fail
                elif header == "dividend_yield_pct" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 0.5 else fill_fail
                    
            elif preset_name == "Growth Accelerator":
                if header == "pat_cagr_5yr" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 20.0 else fill_fail
                elif header == "revenue_cagr_5yr" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 15.0 else fill_fail
                elif header == "debt_to_equity" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val <= 2.0 else fill_fail
                    
            elif preset_name == "Dividend Champion":
                if header == "dividend_yield_pct" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 2.0 else fill_fail
                elif header == "dividend_payout_ratio_pct" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val <= 80.0 else fill_fail
                elif header == "free_cash_flow_cr" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val > 0 else fill_fail
                    
            elif preset_name == "Debt-Free Blue Chip":
                if header == "debt_to_equity" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val == 0.0 else fill_fail
                elif header == "return_on_equity_pct" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 12.0 else fill_fail
                elif header == "sales" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 5000.0 else fill_fail
                    
            elif preset_name == "Turnaround Watch":
                if header == "rev_cagr_3yr" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val >= 10.0 else fill_fail
                elif header == "free_cash_flow_cr" and isinstance(val, (int, float)):
                    cell.fill = fill_pass if val > 0 else fill_fail
                    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = ws.cell(row=1, column=col[0].column).column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def evaluate_presets():
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    db_path = os.path.join(base_dir, "data", "nifty100.db")
    output_path = os.path.join(base_dir, "output", "screener_output.xlsx")
    
    # Initialize engine
    engine = ScreenerEngine(db_path=db_path)
    df = engine.get_latest_data()
    
    # Re-calculate or fallback to sector relative composite score
    df = compute_sector_relative_scores(df)
    
    # Calculate 3yr CAGR and D/E trends for Turnaround Watch
    extra_trends = get_3yr_rev_cagr_and_de_trend(db_path)
    df['rev_cagr_3yr'] = df['company_id'].apply(lambda x: extra_trends.get(x, {}).get('rev_cagr_3yr'))
    df['de_declining'] = df['company_id'].apply(lambda x: extra_trends.get(x, {}).get('de_declining'))
    
    # Define the 6 Presets
    presets = {
        "Quality Compounder": {
            "min_roe": 15.0,
            "max_de": 1.0,
            "min_fcf": 0.0,
            "min_rev_cagr_5yr": 10.0
        },
        "Value Pick": {
            "max_pe": 30.0,
            "max_pb": 5.0,
            "max_de": 2.0,
            "min_dividend_yield": 0.5
        },
        "Growth Accelerator": {
            "min_pat_cagr_5yr": 20.0,
            "min_rev_cagr_5yr": 15.0,
            "max_de": 2.0
        },
        "Dividend Champion": {
            "min_dividend_yield": 2.0,
            "max_dividend_payout": 80.0,
            "min_fcf": 0.0
        },
        "Debt-Free Blue Chip": {
            "max_de": 0.0,
            "min_roe": 12.0,
            "min_sales": 5000.0
        },
        "Turnaround Watch": {
            # Turnaround Watch logic handles custom boolean or calculations
            # Applied manually below
        }
    }
    
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for name, criteria in presets.items():
        if name == "Turnaround Watch":
            # Handled manually
            mask = (df['rev_cagr_3yr'].fillna(-999.0) >= 10.0) & (df['free_cash_flow_cr'].fillna(-999.0) > 0) & (df['de_declining'] == True)
            res_df = df[mask].sort_values(by="sector_composite_quality_score", ascending=False)
        else:
            res_df = engine.apply_filters(df, criteria)
            # Re-sort by sector relative composite score
            res_df = res_df.sort_values(by="sector_composite_quality_score", ascending=False)
            
        print(f"Preset {name} returns: {len(res_df)} companies.")
        
        # Clean helper columns for display
        display_df = res_df.copy()
        if 'de_declining' in display_df.columns:
            display_df = display_df.drop(columns=['de_declining'])
            
        ws = wb.create_sheet(title=name)
        
        # Write header
        ws.append(list(display_df.columns))
        # Write data
        for row in display_df.itertuples(index=False):
            ws.append(list(row))
            
        apply_excel_styling(ws, name)
        
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Presets evaluation complete. Saved to {output_path}")

if __name__ == "__main__":
    evaluate_presets()
